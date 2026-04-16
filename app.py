# app.py — Payroll System Web Application

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, session
from functools import wraps
from models import get_db, init_db
from payroll_calc import calculate_employee_payroll, run_monthly_payroll, save_payroll
from preview_calc import calculate_preview
from finger_import import import_and_save
from excel_import import import_payroll_excel
from sid_import import import_sid_file, STATUS_MAP
from leave_calc import init_leave_balance, use_leave, cancel_leave, get_leave_summary, get_employee_leave_detail
from config import PTKP
import pandas as pd
import os
from datetime import datetime, date, timedelta

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'payroll-nadia-2026')
app.permanent_session_lifetime = timedelta(hours=8)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'data', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ============================================================
# Authentication
# ============================================================
USERS = {
    'admin': os.environ.get('ADMIN_PASSWORD', 'komi2026!'),
    'nadia': os.environ.get('NADIA_PASSWORD', 'nadia2026!'),
    'hrd': os.environ.get('HRD_PASSWORD', 'hrd2026!'),
}

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        password = request.form.get('password', '')
        if username in USERS and USERS[username] == password:
            session.permanent = True
            session['user'] = username
            flash(f'Selamat datang, {username}!', 'success')
            next_url = request.args.get('next', url_for('dashboard'))
            return redirect(next_url)
        else:
            flash('Username atau password salah!', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('user', None)
    flash('Berhasil logout.', 'info')
    return redirect(url_for('login'))

def format_rupiah(amount):
    """Format number as Rupiah."""
    if amount is None:
        return "Rp 0"
    try:
        return f"Rp {amount:,.0f}".replace(",", ".")
    except (ValueError, TypeError):
        return "Rp 0"

app.jinja_env.filters['rupiah'] = format_rupiah

def get_factories():
    conn = get_db()
    factories = conn.execute('SELECT * FROM factories WHERE is_active = 1 ORDER BY code').fetchall()
    conn.close()
    return factories

# ============================================================
# Dashboard
# ============================================================
@app.route('/')
@login_required
def dashboard():
    conn = get_db()
    factory_id = request.args.get('factory', '')
    
    factory_filter = ''
    params = []
    if factory_id:
        factory_filter = ' AND e.factory_id = ?'
        params = [int(factory_id)]
    
    total_employees = conn.execute(
        f'SELECT COUNT(*) FROM employees e WHERE e.is_active = 1{factory_filter}', params
    ).fetchone()[0]
    
    current_period = date.today().strftime('%Y-%m')
    payroll_done = conn.execute(
        f'''SELECT COUNT(*) FROM payroll p 
            JOIN employees e ON p.employee_id = e.id 
            WHERE p.period = ?{factory_filter}''', 
        [current_period] + params
    ).fetchone()[0]
    
    total_payroll = conn.execute(
        f'''SELECT SUM(p.net_salary) FROM payroll p 
            JOIN employees e ON p.employee_id = e.id 
            WHERE p.period = ?{factory_filter}''',
        [current_period] + params
    ).fetchone()[0] or 0
    
    total_gross = conn.execute(
        f'''SELECT SUM(p.gross_salary) FROM payroll p 
            JOIN employees e ON p.employee_id = e.id 
            WHERE p.period = ?{factory_filter}''',
        [current_period] + params
    ).fetchone()[0] or 0
    
    total_overtime = conn.execute(
        f'''SELECT SUM(p.overtime_pay) FROM payroll p 
            JOIN employees e ON p.employee_id = e.id 
            WHERE p.period = ?{factory_filter}''',
        [current_period] + params
    ).fetchone()[0] or 0
    
    recent_payrolls = conn.execute(f'''
        SELECT p.*, e.name, e.nik, f.code as factory_code FROM payroll p 
        JOIN employees e ON p.employee_id = e.id 
        LEFT JOIN factories f ON e.factory_id = f.id
        WHERE 1=1 {factory_filter.replace('AND e.', 'AND e.')}
        ORDER BY p.created_at DESC LIMIT 10
    ''', params).fetchall()
    
    # Per-factory summary
    factory_stats = conn.execute('''
        SELECT f.code, f.name, COUNT(e.id) as emp_count
        FROM factories f LEFT JOIN employees e ON f.id = e.factory_id AND e.is_active = 1
        WHERE f.is_active = 1
        GROUP BY f.id ORDER BY f.code
    ''').fetchall()
    
    # Available periods
    periods = conn.execute(
        'SELECT DISTINCT period FROM payroll ORDER BY period DESC LIMIT 12'
    ).fetchall()

    # --- OT data for dashboard ---
    today_str = date.today().isoformat()
    daily_ot_cost = conn.execute(
        'SELECT SUM(ot_cost) FROM overtime_requests WHERE request_date = ?', (today_str,)
    ).fetchone()[0] or 0

    # Weekly OT cost (Mon-Fri)
    week_dates = _current_week_dates()
    holidays = _get_holidays_set(conn)
    weekly_ot = []
    for d in week_dates:
        if d in holidays:
            weekly_ot.append({'date': d, 'cost': 0, 'label': d[-5:], 'holiday': True})
        else:
            c = conn.execute(
                'SELECT SUM(ot_cost) FROM overtime_requests WHERE request_date = ?', (d,)
            ).fetchone()[0] or 0
            weekly_ot.append({'date': d, 'cost': c, 'label': d[-5:], 'holiday': False})

    # OT Warning: employees > 15 hours this week (Mon-Fri excl holidays)
    valid_dates = [d for d in week_dates if d not in holidays]
    ot_warning_employees = []
    if valid_dates:
        placeholders = ','.join('?' * len(valid_dates))
        ot_warning_employees = conn.execute(f'''
            SELECT e.nik, e.name, e.department, SUM(ot.actual_hours) as total_hours
            FROM overtime_requests ot
            JOIN employees e ON ot.employee_id = e.id
            WHERE ot.request_date IN ({placeholders})
            GROUP BY e.nik, e.name, e.department
            HAVING SUM(ot.actual_hours) > 15
            ORDER BY SUM(ot.actual_hours) DESC
        ''', valid_dates).fetchall()

    conn.close()
    return render_template('dashboard.html',
        total_employees=total_employees,
        payroll_done=payroll_done,
        total_payroll=total_payroll,
        total_gross=total_gross,
        total_overtime=total_overtime,
        current_period=current_period,
        recent_payrolls=recent_payrolls,
        factories=get_factories(),
        factory_stats=factory_stats,
        selected_factory=factory_id,
        periods=periods,
        daily_ot_cost=daily_ot_cost,
        weekly_ot=weekly_ot,
        ot_warning_employees=ot_warning_employees
    )

# ============================================================
# Employees (Karyawan)
# ============================================================
@app.route('/employees')
@login_required
def employees():
    conn = get_db()
    factory_id = request.args.get('factory', '')
    search = request.args.get('search', '').strip()
    
    query = '''
        SELECT e.*, f.code as factory_code FROM employees e
        LEFT JOIN factories f ON e.factory_id = f.id
        WHERE 1=1
    '''
    params = []
    
    if factory_id:
        query += ' AND e.factory_id = ?'
        params.append(int(factory_id))
    
    if search:
        query += ' AND (e.name LIKE ? OR e.nik LIKE ?)'
        params.extend([f'%{search}%', f'%{search}%'])
    
    query += ' ORDER BY f.code, e.name'
    
    emps = conn.execute(query, params).fetchall()
    conn.close()
    return render_template('employees.html', employees=emps, 
                         factories=get_factories(), selected_factory=factory_id, search=search)

@app.route('/employees/add', methods=['GET', 'POST'])
@login_required
def add_employee():
    if request.method == 'POST':
        conn = get_db()
        try:
            conn.execute('''
                INSERT INTO employees (nik, name, factory_id, department, position, section, join_date,
                    base_salary, transport_allowance, meal_allowance, ptkp_status,
                    bank_name, bank_account, finger_id, work_schedule)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (
                request.form['nik'], request.form['name'],
                int(request.form.get('factory_id', 1)),
                request.form.get('department', ''), request.form.get('position', ''),
                request.form.get('section', ''),
                request.form['join_date'], float(request.form['base_salary']),
                float(request.form.get('transport_allowance', 0)),
                float(request.form.get('meal_allowance', 0)),
                request.form.get('ptkp_status', 'TK/0'),
                request.form.get('bank_name', ''), request.form.get('bank_account', ''),
                request.form.get('finger_id', ''),
                request.form.get('work_schedule', '')
            ))
            conn.commit()
            flash('Karyawan berhasil ditambahkan!', 'success')
        except Exception as e:
            flash(f'Error: {e}', 'danger')
        finally:
            conn.close()
        return redirect(url_for('employees'))
    
    return render_template('employee_form.html', employee=None, 
                         ptkp_options=list(PTKP.keys()), factories=get_factories())

@app.route('/employees/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_employee(id):
    conn = get_db()
    if request.method == 'POST':
        try:
            conn.execute('''
                UPDATE employees SET nik=?, name=?, factory_id=?, department=?, position=?,
                    section=?, join_date=?, base_salary=?, transport_allowance=?, meal_allowance=?,
                    ptkp_status=?, bank_name=?, bank_account=?, finger_id=?, work_schedule=?,
                    updated_at=CURRENT_TIMESTAMP
                WHERE id=?
            ''', (
                request.form['nik'], request.form['name'],
                int(request.form.get('factory_id', 1)),
                request.form.get('department', ''), request.form.get('position', ''),
                request.form.get('section', ''),
                request.form['join_date'], float(request.form['base_salary']),
                float(request.form.get('transport_allowance', 0)),
                float(request.form.get('meal_allowance', 0)),
                request.form.get('ptkp_status', 'TK/0'),
                request.form.get('bank_name', ''), request.form.get('bank_account', ''),
                request.form.get('finger_id', ''),
                request.form.get('work_schedule', ''), id
            ))
            conn.commit()
            flash('Data karyawan berhasil diupdate!', 'success')
        except Exception as e:
            flash(f'Error: {e}', 'danger')
        finally:
            conn.close()
        return redirect(url_for('employees'))
    
    emp = conn.execute('SELECT * FROM employees WHERE id = ?', (id,)).fetchone()
    conn.close()
    return render_template('employee_form.html', employee=emp, 
                         ptkp_options=list(PTKP.keys()), factories=get_factories())

@app.route('/employees/import', methods=['GET', 'POST'])
@login_required
def import_employees():
    """Bulk import employees from Excel/CSV."""
    if request.method == 'POST':
        file = request.files.get('file')
        if not file:
            flash('Pilih file dulu!', 'danger')
            return redirect(url_for('import_employees'))
        
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)
        
        try:
            if filepath.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(filepath)
            else:
                df = pd.read_csv(filepath)
            
            df.columns = [c.strip().lower().replace(' ', '_') for c in df.columns]
            
            conn = get_db()
            saved = 0
            skipped = 0
            
            for _, row in df.iterrows():
                try:
                    nik = str(row.get('nik', '')).strip()
                    name = str(row.get('name', row.get('nama', ''))).strip()
                    if not nik or not name:
                        skipped += 1
                        continue
                    
                    factory_code = str(row.get('factory', row.get('pabrik', 'F1'))).strip()
                    factory = conn.execute('SELECT id FROM factories WHERE code = ?', (factory_code,)).fetchone()
                    factory_id = factory['id'] if factory else 1
                    
                    conn.execute('''
                        INSERT OR IGNORE INTO employees 
                        (nik, name, factory_id, department, position, join_date,
                         base_salary, transport_allowance, meal_allowance, ptkp_status,
                         bank_name, bank_account, finger_id)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                    ''', (
                        nik, name, factory_id,
                        str(row.get('department', row.get('departemen', ''))),
                        str(row.get('position', row.get('jabatan', ''))),
                        str(row.get('join_date', row.get('tanggal_masuk', '2025-01-01'))),
                        float(row.get('base_salary', row.get('gaji_pokok', 0))),
                        float(row.get('transport_allowance', row.get('transport', 0))),
                        float(row.get('meal_allowance', row.get('makan', 0))),
                        str(row.get('ptkp_status', row.get('ptkp', 'TK/0'))),
                        str(row.get('bank_name', row.get('bank', ''))),
                        str(row.get('bank_account', row.get('rekening', ''))),
                        str(row.get('finger_id', ''))
                    ))
                    saved += 1
                except Exception:
                    skipped += 1
            
            conn.commit()
            conn.close()
            flash(f'Import selesai! {saved} karyawan ditambahkan, {skipped} dilewati.', 'success')
        except Exception as e:
            flash(f'Error: {e}', 'danger')
        
        return redirect(url_for('employees'))
    
    return render_template('import_employees.html')

@app.route('/employees/template')
@login_required
def employee_template():
    """Download Excel template for bulk import."""
    df = pd.DataFrame({
        'nik': ['001', '002'],
        'name': ['Budi Santoso', 'Siti Aminah'],
        'factory': ['F1', 'F2'],
        'department': ['Produksi', 'QC'],
        'position': ['Operator', 'Inspector'],
        'join_date': ['2024-01-15', '2023-06-01'],
        'base_salary': [5000000, 5500000],
        'transport_allowance': [500000, 500000],
        'meal_allowance': [400000, 400000],
        'ptkp_status': ['TK/0', 'K/1'],
        'bank_name': ['BCA', 'Mandiri'],
        'bank_account': ['1234567890', '0987654321'],
        'finger_id': ['F001', 'F002'],
    })
    export_dir = os.path.join(os.path.dirname(__file__), 'exports')
    os.makedirs(export_dir, exist_ok=True)
    path = os.path.join(export_dir, 'template_karyawan.xlsx')
    df.to_excel(path, index=False)
    return send_file(path, as_attachment=True)

# ============================================================
# Attendance (Absensi)
# ============================================================
@app.route('/attendance')
@login_required
def attendance():
    period = request.args.get('period', date.today().strftime('%Y-%m'))
    conn = get_db()
    records = conn.execute('''
        SELECT a.*, e.name, e.nik FROM attendance a
        JOIN employees e ON a.employee_id = e.id
        WHERE a.date LIKE ?
        ORDER BY a.date DESC, e.name
    ''', (f'{period}%',)).fetchall()
    conn.close()
    return render_template('attendance.html', records=records, period=period)

# ============================================================
# SID (Status Ijin/Dispensasi) Import
# ============================================================
@app.route('/attendance/sid', methods=['GET', 'POST'])
@login_required
def import_sid():
    """Import SID template — attendance status (ijin, sakit, alpha, etc.)."""
    current_period = request.args.get('period', date.today().strftime('%Y-%m'))
    
    if request.method == 'POST':
        file = request.files.get('file')
        period = request.form.get('period', current_period)
        
        if not file:
            flash('Pilih file SID template dulu!', 'danger')
            return redirect(url_for('import_sid'))
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('File harus format Excel (.xlsx)!', 'danger')
            return redirect(url_for('import_sid'))
        
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)
        
        try:
            stats = import_sid_file(filepath, period)
            
            status_detail = ', '.join(
                f'{code}: {count}건' for code, count in stats['by_status'].items()
            ) if stats['by_status'] else '없음'
            
            msg = (
                f"SID Import 완료! 기간: {period}<br>"
                f"✅ 업데이트: {stats['updated']}건<br>"
                f"⏭️ 건너뜀: {stats['skipped']}건<br>"
                f"📊 상태별: {status_detail}"
            )
            if stats['errors'][:5]:
                error_list = '<br>'.join(stats['errors'][:5])
                if len(stats['errors']) > 5:
                    error_list += f'<br>... 외 {len(stats["errors"])-5}건'
                msg += f"<br>⚠️ 오류:<br>{error_list}"
            
            flash(msg, 'success')
            return redirect(url_for('attendance', period=period))
            
        except Exception as e:
            flash(f'Error import SID: {e}', 'danger')
            return redirect(url_for('import_sid'))
    
    return render_template('import_sid.html', current_period=current_period)


@app.route('/attendance/sid/template')
@login_required
def download_sid_template():
    """Generate and download SID template for a given period."""
    period = request.args.get('period', date.today().strftime('%Y-%m'))
    parts = period.split('-')
    year, month = int(parts[0]), int(parts[1])
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    import calendar
    
    max_day = calendar.monthrange(year, month)[1]
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SID_Template"
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_day+3)
    title_cell = ws.cell(row=1, column=1, 
        value=f"SID TEMPLATE - {calendar.month_name[month].upper()} {year} ({month}월)")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ["NIK", "NAME", "FACTORY"] + [str(d) for d in range(1, max_day+1)]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Employee data
    conn = get_db()
    employees = conn.execute("""
        SELECT nik, name, 
               CASE factory_id WHEN 1 THEN 'PUSAT' WHEN 2 THEN 'WONSA' END as factory
        FROM employees WHERE is_active = 1 ORDER BY factory_id, name
    """).fetchall()
    conn.close()
    
    # Data validation
    dv = DataValidation(
        type="list",
        formula1='"H,I,S,A,C,CB,DL,CT,OFF"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid Code",
        error="Pilih: H, I, S, A, C, CB, DL, CT, OFF"
    )
    
    for row_idx, emp in enumerate(employees, 3):
        ws.cell(row=row_idx, column=1, value=emp['nik']).border = thin_border
        ws.cell(row=row_idx, column=2, value=emp['name']).border = thin_border
        ws.cell(row=row_idx, column=3, value=emp['factory']).border = thin_border
        for col in range(4, max_day+4):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
    
    last_row = len(employees) + 2
    dv.add(f"D3:AH{last_row}")
    ws.add_data_validation(dv)
    
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10
    for col in range(4, max_day+4):
        ws.column_dimensions[get_column_letter(col)].width = 5
    
    ws.freeze_panes = 'D3'
    
    # Save
    export_dir = os.path.join(os.path.dirname(__file__), 'exports')
    os.makedirs(export_dir, exist_ok=True)
    filepath = os.path.join(export_dir, f'sid_template_{period.replace("-","")}.xlsx')
    wb.save(filepath)
    
    return send_file(filepath, as_attachment=True, 
                     download_name=f'sid_template_{period.replace("-","")}.xlsx')


@app.route('/attendance/import', methods=['GET', 'POST'])
@login_required
def import_attendance():
    if request.method == 'POST':
        file = request.files.get('file')
        if not file:
            flash('Pilih file dulu!', 'danger')
            return redirect(url_for('import_attendance'))
        
        format_type = request.form.get('format_type', 'generic')
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)
        
        try:
            result = import_and_save(filepath, format_type)
            flash(
                f"Import selesai! {result['saved']} record tersimpan, "
                f"{result['skipped']} dilewati, "
                f"dari total {result['total_records']} scan.",
                'success'
            )
        except Exception as e:
            flash(f'Error import: {e}', 'danger')
        
        return redirect(url_for('attendance'))
    
    return render_template('import_attendance.html')

# ============================================================
# Payroll (Penggajian)
# ============================================================
@app.route('/payroll')
@login_required
def payroll():
    period = request.args.get('period', date.today().strftime('%Y-%m'))
    factory_id = request.args.get('factory', '')
    search = request.args.get('search', '').strip()
    conn = get_db()
    
    factory_filter = ''
    search_filter = ''
    params = [period]
    if factory_id:
        factory_filter = ' AND e.factory_id = ?'
        params.append(int(factory_id))
    if search:
        search_filter = ' AND (e.name LIKE ? OR e.nik LIKE ?)'
        params.extend([f'%{search}%', f'%{search}%'])
    
    records = conn.execute(f'''
        SELECT p.*, e.name, e.nik, e.department, e.section, f.code as factory_code FROM payroll p
        JOIN employees e ON p.employee_id = e.id
        LEFT JOIN factories f ON e.factory_id = f.id
        WHERE p.period = ?{factory_filter}{search_filter}
        ORDER BY f.code, e.name
    ''', params).fetchall()
    
    summary = conn.execute(f'''
        SELECT 
            COUNT(*) as total_employees,
            SUM(p.gross_salary) as total_gross,
            SUM(p.overtime_pay) as total_overtime,
            SUM(p.total_deductions) as total_deductions,
            SUM(p.net_salary) as total_net,
            SUM(p.bpjs_total_company) as total_company_bpjs,
            SUM(p.bpjs_total_employee) as total_employee_bpjs,
            SUM(p.bpjs_kes_company) as bpjs_kes_company,
            SUM(p.bpjs_kes_employee) as bpjs_kes_employee,
            SUM(p.bpjs_jht_company) + SUM(p.bpjs_jkk_company) + SUM(p.bpjs_jkm_company) + SUM(p.bpjs_jp_company) as bpjs_tk_company,
            SUM(p.bpjs_jht_employee) + SUM(p.bpjs_jp_employee) as bpjs_tk_employee
        FROM payroll p JOIN employees e ON p.employee_id = e.id
        WHERE p.period = ?{factory_filter}{search_filter}
    ''', params).fetchone()
    
    # Available periods
    periods = conn.execute(
        'SELECT DISTINCT period FROM payroll ORDER BY period DESC'
    ).fetchall()
    
    conn.close()
    return render_template('payroll.html', records=records, period=period, summary=summary,
                         factories=get_factories(), selected_factory=factory_id, periods=periods,
                         search=search)

@app.route('/payroll/run', methods=['POST'])
@login_required
def run_payroll():
    period = request.form.get('period', date.today().strftime('%Y-%m'))
    include_thr = request.form.get('include_thr') == 'on'
    factory_id = request.form.get('factory') or None
    
    try:
        results = run_monthly_payroll(period, include_thr, factory_id=factory_id)
        factory_label = f' (Pabrik {factory_id})' if factory_id else ' (Semua Pabrik)'
        flash(f'Payroll {period}{factory_label} selesai! {len(results)} karyawan diproses.', 'success')
    except Exception as e:
        flash(f'Error: {e}', 'danger')
    
    return redirect(url_for('payroll', period=period))

@app.route('/payroll/import-excel', methods=['GET', 'POST'])
@login_required
def import_payroll_excel_route():
    """Import payroll data from factory Excel file."""
    if request.method == 'POST':
        file = request.files.get('file')
        period = request.form.get('period', date.today().strftime('%Y-%m'))
        factory_id = int(request.form.get('factory_id', 1))
        
        if not file:
            flash('Pilih file Excel dulu!', 'danger')
            return redirect(url_for('import_payroll_excel_route'))
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('File harus berformat Excel (.xlsx)!', 'danger')
            return redirect(url_for('import_payroll_excel_route'))
        
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)
        
        try:
            stats = import_payroll_excel(filepath, period, factory_id)
            
            msg = (
                f"Import berhasil! Periode: {period}<br>"
                f"👤 Karyawan baru: {stats['created_employees']}<br>"
                f"🔄 Karyawan diupdate: {stats['updated_employees']}<br>"
                f"💰 Record payroll: {stats['payroll_records']}<br>"
                f"⏭️ Dilewati: {stats['skipped']}"
            )
            if stats['errors']:
                msg += f"<br>⚠️ Error: {len(stats['errors'])} baris"
            
            flash(msg, 'success')
            return redirect(url_for('payroll', period=period))
            
        except Exception as e:
            flash(f'Error import: {e}', 'danger')
            return redirect(url_for('import_payroll_excel_route'))
    
    return render_template('import_payroll_excel.html', factories=get_factories())

@app.route('/payroll/export/<period>')
@login_required
def export_payroll(period):
    conn = get_db()
    records = conn.execute('''
        SELECT e.nik, e.name, e.department, e.section, e.bank_name, e.bank_account,
               p.* FROM payroll p
        JOIN employees e ON p.employee_id = e.id
        WHERE p.period = ?
        ORDER BY e.name
    ''', (period,)).fetchall()
    conn.close()
    
    if not records:
        flash('Tidak ada data payroll untuk period ini.', 'warning')
        return redirect(url_for('payroll', period=period))
    
    df = pd.DataFrame([dict(r) for r in records])
    export_dir = os.path.join(os.path.dirname(__file__), 'exports')
    os.makedirs(export_dir, exist_ok=True)
    export_path = os.path.join(export_dir, f'payroll_{period}.xlsx')
    df.to_excel(export_path, index=False)
    
    return send_file(export_path, as_attachment=True)

@app.route('/payroll/slip-all/<period>')
@login_required
def payroll_slip_all(period):
    factory_id = request.args.get('factory', '')
    conn = get_db()
    
    factory_filter = ''
    params = [period]
    if factory_id:
        factory_filter = ' AND e.factory_id = ?'
        params.append(int(factory_id))
    
    records = conn.execute(f'''
        SELECT p.*, e.name, e.nik, e.department, e.position, e.section,
               e.bank_name, e.bank_account, e.work_schedule, f.code as factory_code
        FROM payroll p
        JOIN employees e ON p.employee_id = e.id
        LEFT JOIN factories f ON e.factory_id = f.id
        WHERE p.period = ?{factory_filter}
        ORDER BY f.code, e.department, e.name
    ''', params).fetchall()
    conn.close()
    
    if not records:
        flash('Tidak ada data payroll untuk period ini.', 'warning')
        return redirect(url_for('payroll', period=period))
    
    return render_template('slip_gaji_all.html', records=records, period=period, factory_id=factory_id)

@app.route('/payroll/slip/<int:payroll_id>')
@login_required
def payroll_slip(payroll_id):
    conn = get_db()
    record = conn.execute('''
        SELECT p.*, e.name, e.nik, e.department, e.position, e.section,
               e.bank_name, e.bank_account, e.work_schedule
        FROM payroll p JOIN employees e ON p.employee_id = e.id
        WHERE p.id = ?
    ''', (payroll_id,)).fetchone()
    conn.close()
    
    if not record:
        flash('Slip gaji tidak ditemukan.', 'warning')
        return redirect(url_for('payroll'))
    
    return render_template('slip_gaji.html', record=record)

# ============================================================
# Payroll Preview (Mid-month)
# ============================================================
@app.route('/payroll/preview')
@login_required
def payroll_preview():
    search = request.args.get('search', '').strip()
    dept_filter = request.args.get('department', '').strip()
    section_filter = request.args.get('section', '').strip()

    summary, employees, departments, sections = calculate_preview(
        factory_id=1,
        period_start='2026-03-01',
        period_end='2026-03-11'
    )

    # Apply filters
    filtered = employees
    if search:
        q = search.lower()
        filtered = [e for e in filtered if q in e['name'].lower() or q in str(e['nik']).lower()]
    if dept_filter:
        filtered = [e for e in filtered if e['department'] == dept_filter]
    if section_filter:
        filtered = [e for e in filtered if e['section'] == section_filter]

    # Recalculate summary for filtered results if filtering
    if search or dept_filter or section_filter:
        filt_summary = {
            'total_employees': len(filtered),
            'total_all_in': sum(1 for r in filtered if r['is_all_in']),
            'total_satpam': sum(1 for r in filtered if r['is_satpam']),
            'total_present_days': sum(r['days_present'] for r in filtered),
            'total_absent_days': sum(r['days_absent'] for r in filtered),
            'total_ot_hours': sum(r['total_ot_hours'] for r in filtered),
            'total_reg_ot': sum(r['total_reg_ot'] for r in filtered),
            'total_hol_ot': sum(r['total_hol_ot'] for r in filtered),
            'total_overtime_pay': sum(r['overtime_pay'] for r in filtered),
            'total_base_salary': sum(r['base_salary'] for r in filtered),
            'total_gross': sum(r['gross_salary'] for r in filtered),
            'total_bpjs_employee': sum(r['bpjs_total_employee'] for r in filtered),
            'total_bpjs_company': sum(r['bpjs_total_company'] for r in filtered),
            'total_deductions': sum(r['total_deductions'] for r in filtered),
            'total_net': sum(r['net_salary'] for r in filtered),
            'total_late_count': sum(r['late_count'] for r in filtered),
            'total_attendance_bonus': sum(r['attendance_bonus'] for r in filtered),
            'dept_breakdown': summary['dept_breakdown'],
        }
    else:
        filt_summary = summary

    return render_template('payroll_preview.html',
        summary=filt_summary,
        employees=filtered,
        departments=departments,
        sections=sections,
        search=search,
        selected_dept=dept_filter,
        selected_section=section_filter,
    )

# ============================================================
# Overtime (Lembur) Management
# ============================================================

def _get_holidays_set(conn):
    """Return set of holiday date strings."""
    rows = conn.execute('SELECT date FROM holidays').fetchall()
    return {r['date'] for r in rows}

def _current_week_dates():
    """Return list of Mon-Fri date strings for current week."""
    today = date.today()
    monday = today - timedelta(days=today.weekday())
    return [(monday + timedelta(days=i)).isoformat() for i in range(5)]

def _calc_ot_cost(base_salary, ot_hours, is_holiday=False):
    """
    Calculate OT cost.
    Weekday: first hour 1.5x, subsequent 2x.
    Hourly rate = base_salary / 173.
    """
    if ot_hours <= 0 or base_salary <= 0:
        return 0.0
    hourly = base_salary / 173.0
    if is_holiday:
        # Holiday OT: all hours at 2x (simplified)
        return round(ot_hours * 2.0 * hourly, 0)
    # Weekday
    if ot_hours <= 1:
        return round(ot_hours * 1.5 * hourly, 0)
    return round((1.5 + (ot_hours - 1) * 2.0) * hourly, 0)

def _calc_actual_ot_hours(finger_out_str, normal_end='17:00', dinner_start='19:00', dinner_end='19:30'):
    """
    Calculate actual OT hours from finger_out time.
    Only count hours after normal_end (17:00).
    Deduct dinner 19:00-19:30 if leaves after 20:00.
    """
    if not finger_out_str:
        return 0.0
    try:
        fo = datetime.strptime(finger_out_str, '%H:%M')
        ne = datetime.strptime(normal_end, '%H:%M')
        if fo <= ne:
            return 0.0
        minutes = (fo - ne).total_seconds() / 60.0
        # Deduct dinner if left after 20:00
        if fo >= datetime.strptime('20:00', '%H:%M'):
            minutes -= 30  # 19:00-19:30
        hours = max(0, minutes / 60.0)
        return round(hours, 2)
    except (ValueError, TypeError):
        return 0.0


@app.route('/overtime')
@login_required
def overtime():
    filter_date = request.args.get('date', date.today().isoformat())
    factory_id = request.args.get('factory', '')
    conn = get_db()

    query = '''
        SELECT ot.*, e.nik, e.name, e.department, e.base_salary,
               f.code as factory_code
        FROM overtime_requests ot
        JOIN employees e ON ot.employee_id = e.id
        LEFT JOIN factories f ON ot.factory_id = f.id
        WHERE ot.request_date = ?
    '''
    params = [filter_date]
    if factory_id:
        query += ' AND ot.factory_id = ?'
        params.append(int(factory_id))
    query += ' ORDER BY e.nik'
    records = conn.execute(query, params).fetchall()

    # Summary stats
    total_requests = len(records)
    verified = sum(1 for r in records if r['match_status'] == 'matched')
    mismatch = sum(1 for r in records if r['match_status'] == 'mismatch')
    no_finger = sum(1 for r in records if r['match_status'] == 'no_finger')
    total_cost = sum(r['ot_cost'] for r in records)

    conn.close()
    return render_template('overtime.html',
        records=records,
        filter_date=filter_date,
        factories=get_factories(),
        selected_factory=factory_id,
        total_requests=total_requests,
        verified=verified,
        mismatch=mismatch,
        no_finger=no_finger,
        total_cost=total_cost
    )


@app.route('/overtime/import', methods=['GET', 'POST'])
@login_required
def import_overtime():
    if request.method == 'POST':
        file = request.files.get('file')
        ot_date = request.form.get('ot_date', date.today().isoformat())
        factory_id = int(request.form.get('factory_id', 1))

        if not file:
            flash('Pilih file Excel dulu!', 'danger')
            return redirect(url_for('import_overtime'))

        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)

        try:
            df = pd.read_excel(filepath)
            df.columns = [c.strip() for c in df.columns]

            # Normalise column names
            col_map = {}
            for c in df.columns:
                cl = c.lower()
                if 'nik' in cl:
                    col_map['nik'] = c
                elif 'mulai' in cl or 'start' in cl:
                    col_map['start'] = c
                elif 'selesai' in cl or 'end' in cl:
                    col_map['end'] = c
                elif 'alasan' in cl or 'reason' in cl:
                    col_map['reason'] = c

            conn = get_db()
            saved = 0
            skipped = 0

            for _, row in df.iterrows():
                nik = str(row.get(col_map.get('nik', 'NIK'), '')).strip()
                if not nik:
                    skipped += 1
                    continue

                emp = conn.execute('SELECT id FROM employees WHERE nik = ?', (nik,)).fetchone()
                if not emp:
                    skipped += 1
                    continue

                start_raw = str(row.get(col_map.get('start', 'Jam Mulai'), '17:00')).strip()
                end_raw = str(row.get(col_map.get('end', 'Jam Selesai'), '20:00')).strip()
                reason = str(row.get(col_map.get('reason', 'Alasan'), '')).strip()

                # Parse times — handle HH:MM or datetime
                def parse_time(val):
                    v = str(val).strip()
                    if len(v) >= 5:
                        return v[:5]
                    return v

                p_start = parse_time(start_raw)
                p_end = parse_time(end_raw)

                # Calculate planned hours
                try:
                    t1 = datetime.strptime(p_start, '%H:%M')
                    t2 = datetime.strptime(p_end, '%H:%M')
                    planned_h = max(0, (t2 - t1).total_seconds() / 3600.0)
                except Exception:
                    planned_h = 0

                try:
                    conn.execute('''
                        INSERT OR REPLACE INTO overtime_requests
                        (employee_id, request_date, planned_start, planned_end,
                         planned_hours, reason, factory_id, status)
                        VALUES (?,?,?,?,?,?,?, 'pending')
                    ''', (emp['id'], ot_date, p_start, p_end, round(planned_h, 2),
                          reason, factory_id))
                    saved += 1
                except Exception:
                    skipped += 1

            conn.commit()
            conn.close()
            flash(f'Import lembur selesai! {saved} record tersimpan, {skipped} dilewati.', 'success')
            return redirect(url_for('overtime', date=ot_date))

        except Exception as e:
            flash(f'Error import: {e}', 'danger')
            return redirect(url_for('import_overtime'))

    return render_template('import_overtime.html', factories=get_factories())


@app.route('/overtime/match', methods=['POST'])
@login_required
def match_overtime():
    match_date = request.form.get('match_date', date.today().isoformat())
    conn = get_db()
    holidays = _get_holidays_set(conn)
    is_holiday = match_date in holidays

    ot_requests = conn.execute('''
        SELECT ot.id, ot.employee_id, ot.planned_start, ot.planned_end,
               e.base_salary
        FROM overtime_requests ot
        JOIN employees e ON ot.employee_id = e.id
        WHERE ot.request_date = ?
    ''', (match_date,)).fetchall()

    matched_count = 0
    for ot in ot_requests:
        att = conn.execute('''
            SELECT clock_in, clock_out FROM attendance
            WHERE employee_id = ? AND date = ?
        ''', (ot['employee_id'], match_date)).fetchone()

        if not att or not att['clock_out']:
            conn.execute('''
                UPDATE overtime_requests SET match_status='no_finger',
                    status='pending', finger_in=NULL, finger_out=NULL,
                    actual_hours=0, ot_cost=0
                WHERE id=?
            ''', (ot['id'],))
            continue

        finger_in = att['clock_in'] or ''
        finger_out = att['clock_out'] or ''

        # Normalise to HH:MM
        fi_short = finger_in[:5] if len(finger_in) >= 5 else finger_in
        fo_short = finger_out[:5] if len(finger_out) >= 5 else finger_out

        actual_h = _calc_actual_ot_hours(fo_short)
        cost = _calc_ot_cost(ot['base_salary'] or 0, actual_h, is_holiday)

        planned_end = ot['planned_end'] or '20:00'
        if fo_short >= planned_end:
            status = 'verified'
            m_status = 'matched'
        else:
            status = 'mismatch'
            m_status = 'mismatch'

        conn.execute('''
            UPDATE overtime_requests SET
                finger_in=?, finger_out=?, actual_hours=?, ot_cost=?,
                status=?, match_status=?
            WHERE id=?
        ''', (fi_short, fo_short, actual_h, cost, status, m_status, ot['id']))
        matched_count += 1

    conn.commit()
    conn.close()
    flash(f'Matching selesai! {matched_count} record diproses untuk tanggal {match_date}.', 'success')
    return redirect(url_for('overtime', date=match_date))


@app.route('/overtime/dashboard')
@login_required
def overtime_dashboard():
    return redirect(url_for('dashboard'))


@app.route('/api/overtime-daily')
@login_required
def api_overtime_daily():
    """Return daily OT cost for current week (Mon-Fri)."""
    conn = get_db()
    week_dates = _current_week_dates()
    holidays = _get_holidays_set(conn)
    result = []
    for d in week_dates:
        if d in holidays:
            result.append({'date': d, 'cost': 0, 'holiday': True})
            continue
        row = conn.execute(
            'SELECT SUM(ot_cost) as total FROM overtime_requests WHERE request_date = ?', (d,)
        ).fetchone()
        result.append({'date': d, 'cost': row['total'] or 0, 'holiday': False})
    conn.close()
    return jsonify(result)


# ============================================================
# ============================================================
# Leave (Cuti) Management
# ============================================================
@app.route('/leave')
@login_required
def leave_page():
    selected_year = request.args.get('year', date.today().year, type=int)
    selected_factory = request.args.get('factory', '')
    search = request.args.get('search', '')

    records = get_leave_summary(selected_year, selected_factory or None)

    # Filter by search
    if search:
        s = search.lower()
        records = [r for r in records if s in str(r['nik']).lower() or s in str(r['name']).lower()]

    # Summary stats
    total = len(records)
    eligible = sum(1 for r in records if r['annual_quota'] > 0)
    not_eligible = total - eligible

    conn = get_db()
    cb_days = conn.execute(
        "SELECT COUNT(*) FROM holidays WHERE is_cuti_bersama = 1 AND date LIKE ?",
        (f"{selected_year}-%",)
    ).fetchone()[0]
    conn.close()

    summary = {'total': total, 'eligible': eligible, 'not_eligible': not_eligible, 'cb_days': cb_days}
    negative_balance = [r for r in records if r['remaining'] < 0]
    years = list(range(2024, date.today().year + 2))

    return render_template('leave.html',
        records=records, summary=summary, negative_balance=negative_balance,
        years=years, selected_year=selected_year,
        selected_factory=selected_factory, search=search,
        factories=get_factories())

@app.route('/leave/init', methods=['POST'])
@login_required
def leave_init():
    year = request.form.get('year', date.today().year, type=int)
    stats = init_leave_balance(year)
    flash(f"✅ Cuti {year} initialized: {stats['created']} created, {stats['updated']} updated, {stats['cb_records']} CB records", 'success')
    return redirect(url_for('leave_page', year=year))

@app.route('/leave/<int:employee_id>')
@login_required
def leave_detail(employee_id):
    year = request.args.get('year', date.today().year, type=int)
    balance, records = get_employee_leave_detail(employee_id, year)
    if not balance:
        flash('⚠️ 해당 직원의 Cuti 데이터가 없습니다. Init을 먼저 실행하세요.', 'warning')
        return redirect(url_for('leave_page', year=year))
    return render_template('leave_detail.html', balance=balance, records=records, year=year)

@app.route('/leave/<int:employee_id>/add', methods=['POST'])
@login_required
def leave_add(employee_id):
    leave_date = request.form['leave_date']
    leave_type = request.form.get('leave_type', 'cuti')
    notes = request.form.get('notes', '')
    year = request.form.get('year', date.today().year, type=int)
    try:
        use_leave(employee_id, leave_date, leave_type, notes)
        flash(f'✅ {leave_date} cuti 등록 완료', 'success')
    except Exception as e:
        flash(f'❌ Error: {e}', 'danger')
    return redirect(url_for('leave_detail', employee_id=employee_id, year=year))

@app.route('/leave/<int:employee_id>/cancel', methods=['POST'])
@login_required
def leave_cancel(employee_id):
    leave_date = request.form['leave_date']
    year = request.form.get('year', date.today().year, type=int)
    try:
        cancel_leave(employee_id, leave_date)
        flash(f'✅ {leave_date} cuti 취소 완료', 'success')
    except Exception as e:
        flash(f'❌ Error: {e}', 'danger')
    return redirect(url_for('leave_detail', employee_id=employee_id, year=year))

# ============================================================
# API endpoints for quick stats
# ============================================================
@app.route('/api/payroll-summary/<period>')
@login_required
def api_payroll_summary(period):
    conn = get_db()
    summary = conn.execute('''
        SELECT 
            COUNT(*) as total_employees,
            SUM(gross_salary) as total_gross,
            SUM(overtime_pay) as total_overtime,
            SUM(net_salary) as total_net,
            SUM(bpjs_total_company) as total_bpjs_company,
            SUM(bpjs_total_employee) as total_bpjs_employee
        FROM payroll WHERE period = ?
    ''', (period,)).fetchone()
    conn.close()
    return jsonify(dict(summary))

# ============================================================
# Finger Lookup — NIK별 출퇴근 조회 & 인쇄
# ============================================================
DAY_NAMES_KR = {0:'월', 1:'화', 2:'수', 3:'목', 4:'금', 5:'토', 6:'일'}
DAY_NAMES_ID = {0:'Sen', 1:'Sel', 2:'Rab', 3:'Kam', 4:'Jum', 5:'Sab', 6:'Min'}

def _get_finger_data(employee_id, month):
    """Get attendance records for an employee for a given month, filling all calendar days."""
    from calendar import monthrange
    conn = get_db()
    
    year, mon = int(month.split('-')[0]), int(month.split('-')[1])
    days_in_month = monthrange(year, mon)[1]
    
    # Get holidays
    holidays_rows = conn.execute(
        "SELECT date FROM holidays WHERE substr(date,1,7) = ?", (month,)
    ).fetchall()
    holiday_dates = set(r['date'] if isinstance(r, dict) else r[0] for r in holidays_rows)
    
    # Get attendance records
    rows = conn.execute('''
        SELECT a.* FROM attendance a
        WHERE a.employee_id = ? AND substr(a.date,1,7) = ?
        ORDER BY a.date
    ''', (employee_id, month)).fetchall()
    conn.close()
    
    att_map = {}
    for r in rows:
        d = r['date'] if isinstance(r, dict) else r[2]
        att_map[d] = r
    
    records = []
    summary = {'present':0, 'late':0, 'absent':0, 'leave':0, 'sid':0, 'total_ot':0.0}
    
    for day in range(1, days_in_month + 1):
        dt = date(year, mon, day)
        date_str = dt.strftime('%Y-%m-%d')
        weekday = dt.weekday()
        day_name = f"{DAY_NAMES_KR[weekday]}/{DAY_NAMES_ID[weekday]}"
        is_holiday = weekday == 6 or date_str in holiday_dates  # Sunday or holiday
        
        rec = att_map.get(date_str)
        if rec:
            clock_in = rec['clock_in'] if isinstance(rec, dict) else rec[3]
            clock_out = rec['clock_out'] if isinstance(rec, dict) else rec[4]
            status = rec['status'] if isinstance(rec, dict) else rec[5]
            ot = rec['overtime_hours'] if isinstance(rec, dict) else rec[6]
            notes = rec['notes'] if isinstance(rec, dict) else rec[8]
            leave_code = rec['leave_code'] if isinstance(rec, dict) else rec[9]
            leave_note = rec['leave_note'] if isinstance(rec, dict) else rec[10]
        else:
            clock_in = clock_out = None
            status = 'holiday' if is_holiday else ''
            ot = 0
            notes = leave_code = leave_note = ''
        
        # Determine display status
        if is_holiday and not rec:
            display_status = 'holiday'
        elif status in ('present', 'hadir'):
            # Check late (after 08:15)
            if clock_in and clock_in > '08:15':
                display_status = 'late'
                summary['late'] += 1
            else:
                display_status = 'present'
            summary['present'] += 1
        elif status in ('alpha', 'absent'):
            display_status = 'alpha'
            summary['absent'] += 1
        elif status in ('leave', 'cuti') or leave_code:
            display_status = 'leave'
            summary['leave'] += 1
        elif status == 'sid':
            display_status = 'sid'
            summary['sid'] += 1
        else:
            display_status = status or ''
        
        if ot:
            summary['total_ot'] += float(ot or 0)
        
        records.append({
            'date': date_str,
            'day_name': day_name,
            'clock_in': clock_in,
            'clock_out': clock_out,
            'status': display_status,
            'overtime_hours': float(ot or 0),
            'notes': notes,
            'leave_code': leave_code,
            'leave_note': leave_note,
            'is_holiday': is_holiday,
        })
    
    return records, summary


@app.route('/finger')
@login_required
def finger_lookup():
    conn = get_db()
    q = request.args.get('q', '').strip()
    selected_month = request.args.get('month', date.today().strftime('%Y-%m'))
    selected_factory = request.args.get('factory', '')
    
    # Available months
    months_rows = conn.execute(
        "SELECT DISTINCT substr(date,1,7) as m FROM attendance ORDER BY m DESC"
    ).fetchall()
    months = [r['m'] if isinstance(r, dict) else r[0] for r in months_rows]
    if not months:
        months = [date.today().strftime('%Y-%m')]
    if selected_month not in months:
        months.insert(0, selected_month)
    
    factories = conn.execute('SELECT * FROM factories WHERE is_active = 1 ORDER BY code').fetchall()
    
    employee = None
    employees = []
    records = None
    summary = None
    
    if q:
        factory_filter = ' AND e.factory_id = ?' if selected_factory else ''
        params_base = [selected_factory] if selected_factory else []
        
        # Try exact NIK match first
        params = [q] + params_base
        emp = conn.execute(f'''
            SELECT e.*, f.code as factory_code FROM employees e
            LEFT JOIN factories f ON e.factory_id = f.id
            WHERE e.nik = ? AND e.is_active = 1 {factory_filter}
        ''', params).fetchone()
        
        if emp:
            employee = emp
        else:
            # Search by name or partial NIK
            params = [f'%{q}%', f'%{q}%'] + params_base
            results = conn.execute(f'''
                SELECT e.*, f.code as factory_code FROM employees e
                LEFT JOIN factories f ON e.factory_id = f.id
                WHERE (e.name LIKE ? OR e.nik LIKE ?) AND e.is_active = 1 {factory_filter}
                ORDER BY e.name LIMIT 50
            ''', params).fetchall()
            
            if len(results) == 1:
                employee = results[0]
            elif len(results) > 1:
                employees = results
    
    conn.close()
    
    if employee:
        emp_id = employee['id'] if isinstance(employee, dict) else employee[0]
        records, summary = _get_finger_data(emp_id, selected_month)
    
    return render_template('finger_lookup.html',
        q=q, employee=employee, employees=employees,
        records=records, summary=summary,
        months=months, selected_month=selected_month,
        factories=factories, selected_factory=selected_factory or '')


@app.route('/finger/print')
@login_required
def finger_print():
    employee_id = request.args.get('employee_id', type=int)
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    
    if not employee_id:
        return "Employee ID required", 400
    
    conn = get_db()
    employee = conn.execute('''
        SELECT e.*, f.code as factory_code FROM employees e
        LEFT JOIN factories f ON e.factory_id = f.id
        WHERE e.id = ?
    ''', (employee_id,)).fetchone()
    conn.close()
    
    if not employee:
        return "Employee not found", 404
    
    records, summary = _get_finger_data(employee_id, month)
    
    return render_template('finger_print.html',
        employee=employee, records=records, summary=summary, month=month)


@app.route('/finger/export')
@login_required
def finger_export():
    employee_id = request.args.get('employee_id', type=int)
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    
    if not employee_id:
        return "Employee ID required", 400
    
    conn = get_db()
    employee = conn.execute('''
        SELECT e.*, f.code as factory_code FROM employees e
        LEFT JOIN factories f ON e.factory_id = f.id
        WHERE e.id = ?
    ''', (employee_id,)).fetchone()
    conn.close()
    
    if not employee:
        return "Employee not found", 404
    
    records, summary = _get_finger_data(employee_id, month)
    
    emp_name = employee['name'] if isinstance(employee, dict) else employee[2]
    emp_nik = employee['nik'] if isinstance(employee, dict) else employee[1]
    
    # Create Excel
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Finger {month}"
    
    thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    hdr_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    
    ws.merge_cells('A1:H1')
    ws['A1'] = f"PT. KOMITRANDO EMPORIO — Laporan Absensi"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A2:H2')
    ws['A2'] = f"NIK: {emp_nik} | Nama: {emp_name} | Periode: {month}"
    ws['A2'].font = Font(size=11)
    
    headers = ['No', 'Tanggal', 'Hari', 'Masuk', 'Pulang', 'Status', 'OT', 'Keterangan']
    for i, h in enumerate(headers):
        c = ws.cell(4, i+1, h)
        c.font = hdr_font; c.fill = hdr_fill; c.border = thin
        c.alignment = Alignment(horizontal='center')
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 6
    ws.column_dimensions['H'].width = 25
    
    alpha_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    late_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    holiday_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    
    for idx, r in enumerate(records):
        row = 5 + idx
        status_text = {'present':'Hadir','late':'Telat','alpha':'Alpha','leave':r.get('leave_code','Cuti'),'sid':'SID','holiday':'Libur'}.get(r['status'], r['status'])
        
        vals = [idx+1, r['date'], r['day_name'], r['clock_in'] or '-', r['clock_out'] or '-',
                status_text, f"{r['overtime_hours']:.1f}" if r['overtime_hours'] else '',
                r['notes'] or r['leave_note'] or '']
        
        fill = None
        if r['status'] == 'alpha': fill = alpha_fill
        elif r['status'] == 'late': fill = late_fill
        elif r['is_holiday']: fill = holiday_fill
        
        for j, v in enumerate(vals):
            c = ws.cell(row, j+1, v)
            c.border = thin
            c.alignment = Alignment(horizontal='center') if j < 7 else Alignment()
            if fill: c.fill = fill
    
    # Summary row
    srow = 5 + len(records) + 1
    ws.merge_cells(f'A{srow}:C{srow}')
    ws.cell(srow, 1, "SUMMARY").font = Font(bold=True)
    ws.cell(srow, 4, f"Hadir: {summary['present']}").font = Font(bold=True)
    ws.cell(srow, 5, f"Telat: {summary['late']}").font = Font(bold=True)
    ws.cell(srow, 6, f"Alpha: {summary['absent']}").font = Font(bold=True)
    ws.cell(srow, 7, f"OT: {summary['total_ot']:.1f}").font = Font(bold=True)
    
    filepath = os.path.join(UPLOAD_FOLDER, f"finger_{emp_nik}_{month}.xlsx")
    wb.save(filepath)
    
    return send_file(filepath, as_attachment=True, 
                     download_name=f"Absensi_{emp_nik}_{emp_name}_{month}.xlsx")


# ============================================================
# Initialize DB on import (for gunicorn/cloud deployment)
init_db()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5050, debug=True)
