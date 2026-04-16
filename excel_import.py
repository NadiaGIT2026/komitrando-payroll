# excel_import.py — Import payroll data from factory Excel format
# Supports PT. KOMITRANDO EMPORIO salary Excel (komi PUSAT sheet)

import openpyxl
from models import get_db

# Column mapping (1-indexed to match openpyxl)
COL = {
    'no':               1,   # A
    'nik':              2,   # B
    'name':             3,   # C
    'department':       4,   # D
    'section':          5,   # E
    'work_schedule':    6,   # F
    'bank_account':     7,   # G
    'active':           8,   # H
    'ot_l1':            9,   # I
    'ot_l2':           10,   # J
    'ot_ll1':          11,   # K
    'ot_ll2':          12,   # L
    'ot_ll3':          13,   # M
    'ot_total_hours':  14,   # N
    'short_hours':     15,   # O
    'work_days':       16,   # P
    'absent_days':     17,   # Q
    'total_work_hours':18,   # R
    'base_salary':     19,   # S
    'special_allowance':20,  # T
    'service_bonus':   21,   # U
    'attendance_bonus': 22,  # V
    'gross_salary':    23,   # W
    'ot_lk1':          24,   # X
    'ot_lk2':          25,   # Y
    'ot_ll1_amt':      26,   # Z
    'ot_ll2_amt':      27,   # AA
    'ot_ll3_amt':      28,   # AB
    'ot_total_pay':    29,   # AC
    'bpjs_kes_co':     30,   # AD
    'bpjs_kes_emp':    31,   # AE
    'bpjs_kes_total':  32,   # AF
    'bpjs_jht_co':     33,   # AG
    'bpjs_jht_emp':    34,   # AH
    'bpjs_jht_total':  35,   # AI
    'bpjs_jkk_co':     36,   # AJ
    'bpjs_jkk_emp':    37,   # AK
    'bpjs_jkk_total':  38,   # AL
    'bpjs_jkm_co':     39,   # AM
    'bpjs_jkm_emp':    40,   # AN
    'bpjs_jkm_total':  41,   # AO
    'bpjs_jp_co':      42,   # AP
    'bpjs_jp_emp':     43,   # AQ
    'bpjs_jp_total':   44,   # AR
    'bpjs_total_co':   45,   # AS
    'bpjs_total_emp':  46,   # AT
    'bpjs_total_total':47,   # AU
    'ded_absence':     48,   # AV
    'ded_late':        49,   # AW
    'ded_total':       50,   # AX
    'pph_allowance':   51,   # AY
    'madome':          52,   # AZ
    'correction_underpay': 53, # BA
    'total_additions': 54,   # BB
    'pph':             55,   # BC
    'correction_salary':56,  # BD
    'total_deductions_manual': 57, # BE
    'net_salary':      58,   # BF
    'bank_method':     59,   # BG
}

def _val(ws, row, col_key, default=0):
    """Get cell value with default."""
    v = ws.cell(row=row, column=COL[col_key]).value
    if v is None:
        return default
    return v

def _num(ws, row, col_key, default=0):
    """Get numeric cell value."""
    v = ws.cell(row=row, column=COL[col_key]).value
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default

def _str(ws, row, col_key, default=''):
    """Get string cell value."""
    v = ws.cell(row=row, column=COL[col_key]).value
    if v is None:
        return default
    return str(v).strip()


def import_payroll_excel(filepath, period, factory_id=1):
    """
    Import a complete payroll month from factory Excel file.
    
    Args:
        filepath: Path to the Excel file
        period: Payroll period in 'YYYY-MM' format
        factory_id: Factory ID (default 1)
    
    Returns:
        dict with counts: created_employees, updated_employees, payroll_records, skipped, errors
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    # Use first sheet (may have trailing spaces)
    ws = wb[wb.sheetnames[0]]
    
    conn = get_db()
    stats = {
        'created_employees': 0,
        'updated_employees': 0,
        'payroll_records': 0,
        'skipped': 0,
        'errors': [],
    }
    
    # Data starts at row 17
    for row in range(17, ws.max_row + 1):
        nik = _str(ws, row, 'nik')
        name = _str(ws, row, 'name')
        active = _str(ws, row, 'active')
        
        # Skip empty/invalid rows and subtotals
        if not nik or not name:
            continue
        if active.lower() not in ('ya', 'yes', '1'):
            stats['skipped'] += 1
            continue
        
        # Clean NIK (remove .0 from Excel number formatting)
        if '.' in nik:
            try:
                nik = str(int(float(nik)))
            except ValueError:
                pass
        
        try:
            department = _str(ws, row, 'department')
            section = _str(ws, row, 'section')
            work_schedule = _str(ws, row, 'work_schedule')
            bank_account = _str(ws, row, 'bank_account')
            base_salary = _num(ws, row, 'base_salary')
            payment_method = _str(ws, row, 'bank_method', 'TRANSFER')
            
            # Ensure/update employee record
            emp = conn.execute('SELECT id FROM employees WHERE nik = ?', (nik,)).fetchone()
            if emp:
                emp_id = emp['id']
                conn.execute('''
                    UPDATE employees SET 
                        name=?, factory_id=?, department=?, section=?,
                        work_schedule=?, bank_account=?, base_salary=?,
                        is_active=1, updated_at=CURRENT_TIMESTAMP
                    WHERE id=?
                ''', (name, factory_id, department, section,
                      work_schedule, bank_account, base_salary, emp_id))
                stats['updated_employees'] += 1
            else:
                cursor = conn.execute('''
                    INSERT INTO employees 
                        (nik, name, factory_id, department, section, join_date,
                         base_salary, work_schedule, bank_account, is_active)
                    VALUES (?,?,?,?,?,?,?,?,?,1)
                ''', (nik, name, factory_id, department, section,
                      '2025-01-01', base_salary, work_schedule, bank_account))
                emp_id = cursor.lastrowid
                stats['created_employees'] += 1
            
            # Insert/replace payroll record with ALL components
            conn.execute('''
                INSERT OR REPLACE INTO payroll (
                    employee_id, period, work_days, absent_days, short_hours, total_work_hours,
                    overtime_l1, overtime_l2, overtime_ll1, overtime_ll2, overtime_ll3, overtime_total_hours,
                    base_salary, special_allowance, service_bonus, attendance_bonus,
                    transport_allowance, meal_allowance, gross_salary,
                    overtime_lk1, overtime_lk2, overtime_ll1_amt, overtime_ll2_amt, overtime_ll3_amt, overtime_pay,
                    bpjs_kes_company, bpjs_kes_employee, bpjs_kes_total,
                    bpjs_jht_company, bpjs_jht_employee, bpjs_jht_total,
                    bpjs_jkk_company, bpjs_jkk_employee, bpjs_jkk_total,
                    bpjs_jkm_company, bpjs_jkm_employee, bpjs_jkm_total,
                    bpjs_jp_company, bpjs_jp_employee, bpjs_jp_total,
                    bpjs_total_company, bpjs_total_employee, bpjs_total_total,
                    deduction_absence, deduction_late, total_deductions,
                    pph_allowance, madome, correction_underpay, total_additions,
                    pph21, correction_salary, total_manual_deductions,
                    net_salary, payment_method, import_source
                ) VALUES (
                    ?,?,?,?,?,?,
                    ?,?,?,?,?,?,
                    ?,?,?,?,
                    ?,?,?,
                    ?,?,?,?,?,?,
                    ?,?,?,
                    ?,?,?,
                    ?,?,?,
                    ?,?,?,
                    ?,?,?,
                    ?,?,?,
                    ?,?,?,
                    ?,?,?,?,
                    ?,?,?,
                    ?,?,?
                )
            ''', (
                emp_id, period,
                int(_num(ws, row, 'work_days')),
                int(_num(ws, row, 'absent_days')),
                _num(ws, row, 'short_hours'),
                _num(ws, row, 'total_work_hours'),
                # Overtime hours
                _num(ws, row, 'ot_l1'),
                _num(ws, row, 'ot_l2'),
                _num(ws, row, 'ot_ll1'),
                _num(ws, row, 'ot_ll2'),
                _num(ws, row, 'ot_ll3'),
                _num(ws, row, 'ot_total_hours'),
                # Earnings
                _num(ws, row, 'base_salary'),
                _num(ws, row, 'special_allowance'),
                _num(ws, row, 'service_bonus'),
                _num(ws, row, 'attendance_bonus'),
                0,  # transport_allowance (not in this Excel)
                0,  # meal_allowance (not in this Excel)
                _num(ws, row, 'gross_salary'),
                # Overtime amounts
                _num(ws, row, 'ot_lk1'),
                _num(ws, row, 'ot_lk2'),
                _num(ws, row, 'ot_ll1_amt'),
                _num(ws, row, 'ot_ll2_amt'),
                _num(ws, row, 'ot_ll3_amt'),
                _num(ws, row, 'ot_total_pay'),
                # BPJS
                _num(ws, row, 'bpjs_kes_co'),
                _num(ws, row, 'bpjs_kes_emp'),
                _num(ws, row, 'bpjs_kes_total'),
                _num(ws, row, 'bpjs_jht_co'),
                _num(ws, row, 'bpjs_jht_emp'),
                _num(ws, row, 'bpjs_jht_total'),
                _num(ws, row, 'bpjs_jkk_co'),
                _num(ws, row, 'bpjs_jkk_emp'),
                _num(ws, row, 'bpjs_jkk_total'),
                _num(ws, row, 'bpjs_jkm_co'),
                _num(ws, row, 'bpjs_jkm_emp'),
                _num(ws, row, 'bpjs_jkm_total'),
                _num(ws, row, 'bpjs_jp_co'),
                _num(ws, row, 'bpjs_jp_emp'),
                _num(ws, row, 'bpjs_jp_total'),
                _num(ws, row, 'bpjs_total_co'),
                _num(ws, row, 'bpjs_total_emp'),
                _num(ws, row, 'bpjs_total_total'),
                # Deductions
                _num(ws, row, 'ded_absence'),
                _num(ws, row, 'ded_late'),
                _num(ws, row, 'ded_total'),
                # Manual additions
                _num(ws, row, 'pph_allowance'),
                _num(ws, row, 'madome'),
                _num(ws, row, 'correction_underpay'),
                _num(ws, row, 'total_additions'),
                # Manual deductions
                _num(ws, row, 'pph'),
                _num(ws, row, 'correction_salary'),
                _num(ws, row, 'total_deductions_manual'),
                # Final
                _num(ws, row, 'net_salary'),
                payment_method if payment_method else 'TRANSFER',
                'excel_import',
            ))
            stats['payroll_records'] += 1
            
        except Exception as e:
            stats['errors'].append(f'Row {row} (NIK {nik}): {str(e)}')
            stats['skipped'] += 1
    
    conn.commit()
    conn.close()
    wb.close()
    
    return stats
